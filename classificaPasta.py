'''
Code by DANIEL PERROUT
03/07/2017
'''

#importing modules
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string
import time
import sys
import datetime
import logging
import sys
from pathlib import Path

#Funcao para coletar todos os elementos de uma coluna do Excel
def lista(cabecalho, coluna):
    titulos = []
    for row in ws.iter_rows(coluna.format(ws.min_row,ws.max_row)):
        for cell in row:
            if cell.value != cabecalho:
                titulos.append(cell.value)
    return titulos

#Funcao que determina a porcentagem de similaridade
def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

print ("##############################################################################")
print ("# Este script faz a sugestão de classificação de acordo com os parâmetros da #")
print ("# Classificação elaborada pela Frente de TI                                  #")
print ("# Versão 1.0 por Daniel Perrout                                              #")
print ("#               <daniel.perrout@planejamento.mg.gov.br>                      #")
print ("##############################################################################")
#COleta o nome do arquivo que deve estar na mesma pasta
nome = (input("Digite o nome do seu arquivo (sem .xlsx):"))
arquivo = ("%s.xlsx" % nome)

#Indicador para continuar
ok = 0
while(ok == 0):
    my_file = Path(arquivo)
    if my_file.is_file():
        print("Arquivo encontrado!")
        ok = 1
    else:
        nome = (input("Arquivo não encontrado. Digite o nome do seu arquivo (sem .xlsx):"))
        arquivo = ("%s.xlsx" % nome)

#Cria o log      
logging.basicConfig(filename=('%s.log' % arquivo), level=logging.INFO, format='%(asctime)s %(message)s')
logging.info('#########################################')
logging.info('# Log para Classificação de Pastas      #')
logging.info('#########################################')    
    
base = "Classificacao.xlsx"

#Abre o arquivo BASE da Proposta
workbook = load_workbook(base)

#Open the first sheet
ws = workbook.active


descricao = lista("DESCRICAO",'A{}:A{}')
proposta = lista("PROPOSTA",'B{}:B{}')
info = "Número de Descrições/Propostas:",len(descricao),"/",len(proposta)
logging.info(info)
print(info)

#Abre o arquivo da Pasta a ser Classificada

workbook = load_workbook(arquivo)

#Open the first sheet
ws = workbook.active

#Coleta Titulos dos Documentos
titulos = lista("TITULO",'A{}:A{}')
classifica = lista("CLASSIFICACAO",'B{}:B{}')
opcoes = []
scores = []
newlist = []

#Percorre todos os TITULOS e classifica
for i, item in enumerate(titulos):
    linha = i+2
    score = 0
    defpro = "N/A"
    numpro = 0
    info = "\nDocumento Página",i+1,":",item
    logging.info(info)
    print(info)
    if item != None:
        for y, desc in enumerate(descricao):
            scoreMax = similar(item,desc)
            scores.append((round(scoreMax*100,2),proposta[y],desc))
            
            if scoreMax > score:
                score = scoreMax
                defpro = desc
                numpro = y
            if item in desc:
                if numpro != y:
                     opcoes.append((proposta[y],desc))

        #Remove duplicados
        if opcoes:
            for i in opcoes:
              if i not in newlist:                  
                newlist.append(i)
  
        
    if defpro == "N/A":
        seq = "N/A"
        opcoes = []
        resultado = "N/A"
    else:
        s = " // ";
        seq = ("{0:.2f}%".format(score*100), "Proposta:", proposta[numpro], "Descrição:", defpro)
        resultado = (s.join( seq ))
    
    

    logging.info((resultado))
    print (resultado)
    
    scores.sort(reverse=True)
    for op in scores[1:4]:     
        info = "    Similares:",op
        logging.info(info)
        print (info)
        
    if newlist:
        ws.cell(row=linha, column=3).value = s.join( newlist[0] )
        for op in newlist:
            info = ("    Ou:",op)
            logging.info(info)
            print (info)
    
    #print ("opcoes:",len(opcoes),len(newlist), len(out))
    
    #print ("Antes:",ws.cell(row=linha, column=2).value)
    ws.cell(row=linha, column=2).value = resultado
    info = ("# Possiveis classificações antes:",len(descricao)," Possiveis classificações agora:",len(newlist)+1)
    logging.info(info)
    print (info)
    opcoes = []
    scores = []
    newlist = []
    
            
#Salva o arquivo
workbook.save(arquivo)

print ("##############################################################################")
print ("# Arquivo \"%s\" salvo com sucesso!" % arquivo )
print ("# Log de operações \"%s.log\" salvo com sucesso" % arquivo )
print ("# Número de páginas classificadas: %s" % len(titulos))
print ("##############################################################################")



 
